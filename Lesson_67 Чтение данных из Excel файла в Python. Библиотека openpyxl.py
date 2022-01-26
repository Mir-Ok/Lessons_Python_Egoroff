''' При работе с файлами есть 3 этапа: открыть файл, обработать, закрыть '''

# для импорта устанавливаем через pip или
# через File | Settings | Project: Уроки Egoroff | Python Interpreter | + справа в углу

import openpyxl

book = openpyxl.open("Pasports.xlsx", read_only=True)  # полный путь, если в другой папке, старый формат xls не читает

sheet = book.active  # метод по умолчанию берет первый лист книги

sheets = book.worksheets
for i in sheets:
    print(i)  # --> <openpyxl.worksheet._read_only.ReadOnlyWorksheet object at 0x000001C80293D5E0> всего 10 в документе

sheet_1 = book.worksheets[1]
print(sheet_1)  # --> <openpyxl.worksheet._read_only.ReadOnlyWorksheet object at 0x000001AE740CC670>
print(sheet_1['A2'])  # --> <ReadOnlyCell 'золото'.A2>
print(sheet_1['A2'].value)  # --> 1

''' Получение значения из ячейки '''

print(sheet['D2'].value)  # --> Студеное, обращение по полному коду D2
print(sheet[2][3].value)  # --> Студеное, обращение к 2 строке (отсчет от 1)
                          # --------------------- и к 3 столбцу (отсчет от 0, по факту он 4-й)

for row in range(1, 11):  # выведем первые 10 строк всех столбцов
    print(sheet[row])  # здесь .value нет

for row in range(1, 11):  # выведем первые 10 строк 4-х столбцов
    col1 = sheet[row][0].value
    col2 = sheet[row][1].value
    col3 = sheet[row][2].value
    col4 = sheet[row][3].value
    print(row, col1, col2, col3, col4)

# если не знаем сколько рядов, то (sheet.max_row +1) вместо 11


''' Работа с диапазонами '''

cells = sheet['B1':'D5']  # диапазон из 3 столбцов
for cell in cells:
    print(cell)  # --> вывел кортеж (<ReadOnlyCell 'золото'.B1>, <ReadOnlyCell 'золото'.C1>, <ReadOnlyCell 'золото'.D1>)

# если кортеж, значит можно использовать множественное присвоение

#for name, info in cells:
#     print(name.value, info.value)
      # --> Рубрикатор | None

for name, info, cell in cells:
      print(name.value, info.value, cell.value)
      # --> Рубрикатор | None | Паспортные характеристики россыпного объекта


''' Работа с диапазонами встроенным методом '''

for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):  # здесь колонки с 1
    print(row)
    # --> (< ReadOnlyCell 'золото'.A1 >, < ReadOnlyCell 'золото'.B1 >, < ReadOnlyCell 'золото'.C1 >)

for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
                           # без уточнения, просто sheet.iter_rows() обойдет весь лист
    for cell in row:  # так как выдает кортеж, по нему тоже можно итерироваться
        print(cell.value, end=' ')
    print()



