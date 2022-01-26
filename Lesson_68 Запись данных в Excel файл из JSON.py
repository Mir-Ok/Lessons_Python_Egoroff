import openpyxl

''' 
book = openpyxl.Workbook()  # создаем рабочую книгу и помещаем ее в переменную

sheet = book.active  # --------- делаем активным первый лист книги
sheet['A2'] = 100  # ----------- изменяем значения выбранных ячеек
sheet['B3'] = 'hello'

sheet[1][0].value = 'world'  # - строки с 1, ячейки с 0, обрщаемся к наполнению
sheet.cell(row=1, column=1).value = 'hello,world'  # перезаписываем значение

book.save('my_book.xlsx')  # --- узакониваем изменения
book.close()  # ---------------- закрываем рабочую книгу

'''

''' Важно! Файл не должен быть открыт другой программой, иначе нет доступа '''

import json

with open('movies.json', encoding='utf-8') as file:
    data = json.load(file)
print(data[' фильмы '])  # в файле старое 'movies' обновлно на ' фильмы '

# for movie in data[' фильмы ']:
#    print(movie)  # получаем вложенные словари, на каждом шаге по одному

for movie in data[' фильмы ']:
    title = movie[' title ']
    year = movie[' год ']
    genres = movie[' жанры ']
    director = movie[' режиссер ']  # это поле пришлось пропарсить на предмет  унификации
    actors = movie[' актеры ']
    print(title, year, genres, director, actors)  # получаем вложенные словари, на каждом шаге по одному

''' 
book = openpyxl.Workbook()  # создаем рабочую книгу и помещаем ее в переменную
sheet = book.active  # --------- делаем активным первый лист книги

sheet['A1'] = 'ID'  # ----------- изменяем значения выбранных ячеек
sheet['B1'] = 'TITLE'  # Ctrl + Shift + U меняет строчные на заглавные
sheet['C1'] = 'YEAR'
sheet['D1'] = 'GENRES'
sheet['E1'] = 'DIRECTOR'
sheet['F1'] = 'ACTORS'

book.save('my_book.xlsx')  # --- узакониваем изменения
book.close()  # ---------------- закрываем рабочую книгу
'''

book = openpyxl.Workbook()  # создаем рабочую книгу и помещаем ее в переменную
sheet = book.active  # --------- делаем активным первый лист книги

sheet['A1'] = 'ID'  # ----------- изменяем значения выбранных ячеек
sheet['B1'] = 'TITLE'  # Ctrl + Shift + U меняет строчные на заглавные
sheet['C1'] = 'YEAR'
sheet['D1'] = 'GENRES'
sheet['E1'] = 'DIRECTOR'
sheet['F1'] = 'ACTORS'

row = 2
for movie in data[' фильмы ']:
    sheet[row][0].value = movie[' id ']
    sheet[row][1].value = movie[' title ']
    sheet[row][2].value = movie[' год ']
    sheet[row][3].value = ' '.join(movie[' жанры '])  # соединим список в строку
    sheet[row][4].value = movie[' режиссер ']  # это поле пришлось пропарсить на предмет  унификации
    sheet[row][5].value = movie[' актеры ']
    row += 1

book.save('my_book.xlsx')  # --- узакониваем изменения
book.close()  # ---------------- закрываем рабочую книгу


